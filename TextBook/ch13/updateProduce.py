import openpyxl

# ワークブックを読み込み、シートを選択
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# 農作物の名前と更新する価格の辞書
PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

# 行を2行目から開始して、ヘッダーをスキップして繰り返します
for row_num in range(2, sheet.max_row + 1):
    # 現在の行の1列目から農作物の名前を取得
    produce_name = sheet.cell(row=row_num, column=1).value
    
    # もし農作物の名前がPRICE_UPDATES辞書にあれば、2列目の価格を更新
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

# 更新したワークブックを保存
wb.save('updatedProduceSales.xlsx')

